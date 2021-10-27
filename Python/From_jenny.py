#-*-coding:utf-8-*-

# Calculate the number of binary files.
 
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

wb = Workbook()
ws = wb.active
ws.title = "Normalization Info"
ws1 = wb.create_sheet("Bill of Material")
#wb.save('JiraReport.xlsx')
wb2 = load_workbook('C:\jira\SAPInnovationManagement.r2sps3.1201.context.xml.Unified.BOM-CTPscan.xlsx')
ws3 = wb2.get_sheet_by_name('Normalization Info')
ws4 = wb2.get_sheet_by_name('Bill of Materials')

max_row = ws3.max_row
max_column = ws3.max_column

for cell in list(ws3.columns)[0]:
    Data = cell.value
   
    if Data is None:
        continue

    str = ''.join(Data)
    if str.startswith( 'Commercial Third-Party Scan Result' ):
        CTPstartrow = cell.row

    if str.startswith( '*GLTS (Global Licensing Ticketing System)' ):
        CTPendrow = cell.row

for row in range(CTPstartrow, CTPstartrow+2):
    for col in range(1, max_column+1):
        cell = ws3.cell(row, col)
        data = cell.value
        if data is None:
            continue

        ws.append([data])

ws.append(['Ticket Number', 'Component', 'Version', 'PPMS SCV Entry', 'PPMS SCV ID', 'Technopedia Release ID'])
A1_font = Font(name='Calibri', size=16, bold=True, underline='single')
A2_font = Font(name='Calibri', size=12, bold=True, color='FF0000')
row3_font = Font(name='Calibri', size=11, bold=True)
row3_alighment = Alignment(horizontal='center', vertical='center')
row_title_fill = PatternFill(fill_type='solid',fgColor='F8CBAD')
row1 = ws.row_dimensions[1]
row2 = ws.row_dimensions[2]
row3 = ws.row_dimensions[3]
ws['A1'].font = A1_font
ws['A2'].font = A2_font
row1.fill = row_title_fill
row2.fill = row_title_fill
ws.cell(row=1, column=1).fill = row_title_fill
ws.cell(row=2, column=1).fill = row_title_fill

for cell in ws["3:3"]:
    cell.font = row3_font
    cell.alignment = row3_alighment

ws.column_dimensions['B'].width = 60
ws.column_dimensions['A'].width = 13.5
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 32
ws.column_dimensions['E'].width = 32
ws.column_dimensions['F'].width = 32

def read_data(ws):
    collect_data=[]
    for row in range(CTPstartrow + 3, CTPendrow - 1):
        sub_data={}
        sub_data['Component']=ws.cell(row,1).value
        sub_data['Version']=ws.cell(row,2).value
        sub_data['Entry']=ws.cell(row,3).value
        sub_data['id']=ws.cell(row,4).value
        sub_data['Technopedia']=ws.cell(row,5).value
        if sub_data['Component'] is not None:
            collect_data.append(sub_data)
    return collect_data

#def read_data(ws):
#    for row in range(CTPstartrow + 3, CTPendrow - 1):
#        for col in range(1, max_column+1):
#           cell = ws.cell(row, col)
#           data2 = cell.value
#           if data2 is None:
#           continue
#   return data2

#ws.append([data2])

def write_data(inputdata, ws):
    row = ws.max_row + 1
    for item in inputdata:
        ws.cell(row, 2).value = item['Component']
        ws.cell(row, 3).value = item['Version']
        ws.cell(row, 4).value = item['Entry'
        ws.cell(row, 5).value = item['id']
        ws.cell(row, 6).value = item['Technopedia']
        row = row + 1

def ws1_info(ws):
    ws["A1"].value = "Commercial Third-Party Scan Result"
    ws.append(['Ticket Number', 'Component', 'Version', 'Component Comment', 'Usage', 'Ship Status', 'Project Source'])
    A1_font = Font(name='Calibri', size=16, bold=True, underline='single')
    row2_font = Font(name='Calibri', size=11, bold=True)
    row2_alighment = Alignment(horizontal='center', vertical='center')
    row1 = ws.row_dimensions[1]
    row2 = ws.row_dimensions[2]
    ws['A1'].font = A1_font
    row1.fill = row_title_fill
    ws.cell(row=1, column=1).fill = row_title_fill
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['A'].width = 13.5
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 32
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 46
    for cell in ws["2:2"]:
        cell.font = row2_font
        cell.alignment = row2_alighment

def read_data2(ws):
    for cell in list(ws.columns)[0]:
        Data2 = cell.value
        if Data2 is None:
            continue

        str2 = ''.join(Data2)
        if str2.startswith( 'Commercial Third-Party Scan Result' ):
            CTPstartrow2 = cell.row
        if str2.startswith( 'Web Service Scan Result' ):
            CTPendrow2 = cell.row
    collect_data1=[]
    for row in range(CTPstartrow2 + 2, CTPendrow2 - 1):
        sub_data={}
        sub_data['Component']=ws.cell(row,1).value
        sub_data['Version']=ws.cell(row,2).value
        sub_data['Comment']=ws.cell(row,3).value
        sub_data['Usage']=ws.cell(row,4).value
        sub_data['Ship']=ws.cell(row,5).value
        sub_data['Source']=ws.cell(row,6).value
        if sub_data['Component'] is not None:
            collect_data1.append(sub_data)

    return collect_data1

 

def write_data2(inputdata2, ws2):
    row = ws2.max_row + 1
    for item in inputdata2:
        ws2.cell(row, 2).value = item['Component']
        ws2.cell(row, 3).value = item['Version']
        ws2.cell(row, 4).value = item['Comment']
        ws2.cell(row, 5).value = item['Usage']
        ws2.cell(row, 6).value = item['Ship']
        ws2.cell(row, 7).value = item['Source']
        row = row + 1

write_data(read_data(ws3), ws)
ws1_info(ws1)
write_data2(read_data2(ws4), ws1)
wb.save('JiraReport.xlsx')

wb.close()
wb2.close()


